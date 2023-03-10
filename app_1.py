
import pymysql.cursors
import openpyxl
from datetime import datetime
from flask import Flask, request, render_template
import os

app = Flask(__name__)

upload_folder='/app/data_2_examine'
app.config['upload_folder'] = upload_folder

def connect_to_db():
    db = pymysql.connect(host='172.17.0.2',
    user='root',         
    passwd='pass', 
    port=3306,
    database='boi')
    return db
          

@app.route("/", methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files['file']
        filename = file.filename
        file.save(os.path.join(app.config['upload_folder'], filename))
        os.chdir(upload_folder)
        for upload_file in os.listdir():
            if upload_file.endswith(".xlsx"):
                wb = openpyxl.load_workbook(upload_file)
                break
            else:
                print("No xlsx files found")
        ws = wb.worksheets[-1]
        db = connect_to_db()
        cursor=db.cursor()
        query1 = ("truncate items;")
        cursor.execute(query1)
        for i in range(2, ws.max_row + 1):#row = [paul.value for paul in ws[i]] #This gets the value of each cell while looping using 'i' as the index
            insertdate = datetime.now()
            date = ws.cell(i,1).value
            full_transaction = ws.cell(i,2).value
            vendor_data = ws.cell(i,2).value.upper()
            v = vendor_data.split(' ')[0]
            if v[0:3] == 'POS':
                vendor=' '.join(vendor_data.split(' ')[1:])
            else:
                vendor = vendor_data
            debit = ws.cell(i,3).value
            if debit is None:
                debit = 0
            credit = ws.cell(i,4).value
            balance = ws.cell(i,5).value
            try:
                query2 = """INSERT INTO items (insertdate,date,full_transaction,vendor,debit,credit,balance) VALUES (%s,%s,%s,%s,%s,%s,%s)"""
                values = (insertdate,date,full_transaction,vendor, debit, credit, balance)
                cursor.execute (query2, values)
                db.commit()
            except:
                print("Unable to connect to db")
                cursor.rollback()
    return render_template('index.html')
  
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8090) 
